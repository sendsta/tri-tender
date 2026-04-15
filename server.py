
from fastmcp import FastMCP
from pathlib import Path
from typing import List, Dict, Any, Optional
import json
import shutil
import pdfplumber
import openpyxl
from datetime import datetime
import re
import zipfile

try:
    import docx  # python-docx
except ImportError:
    docx = None

# Placeholder; we’ll import WeasyPrint lazily inside html_to_pdf
WeasyHTML = None


# ============================================================
# GLOBAL SETUP
# ============================================================

PROJECT_ROOT = Path("projects").absolute()
PROJECT_ROOT.mkdir(parents=True, exist_ok=True)

mcp = FastMCP(
    "TriTender MCP",
    mask_error_details=True,
)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def project_path(project_id: str) -> Path:
    """Resolve path for a project and create if missing."""
    path = PROJECT_ROOT / project_id
    path.mkdir(parents=True, exist_ok=True)
    return path


def safe_write(path: Path, text: str) -> None:
    """Write text to a file, creating parents."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def read_text_file(path: Path, max_chars: int = 200000) -> str:
    """Read a text file, truncated."""
    return path.read_text(encoding="utf-8")[:max_chars]


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


# ============================================================
# BASIC TOOL
# ============================================================

@mcp.tool()
def ping() -> str:
    """Simple health check."""
    return "tri-tender-mcp: OK"


# ============================================================
# SECTION A — INTAKE & STRUCTURING
# ============================================================

@mcp.tool()
def import_tender_pack(project_id: str, source_path: str) -> Dict[str, Any]:
    """
    Import a tender pack from a folder or ZIP into a project's tender_docs directory.
    """
    project = project_path(project_id)
    tender_dir = project / "tender_docs"
    ensure_dir(tender_dir)

    src = Path(source_path).absolute()
    if not src.exists():
        raise ValueError(f"Source path not found: {src}")

    if src.is_file() and src.suffix.lower() == ".zip":
        shutil.unpack_archive(str(src), tender_dir)
    elif src.is_dir():
        for item in src.iterdir():
            if item.is_file():
                shutil.copy(item, tender_dir / item.name)
    else:
        raise ValueError("source_path must be a folder or .zip file")

    imported = [f.name for f in tender_dir.iterdir() if f.is_file()]
    return {
        "status": "ok",
        "project": str(project),
        "imported": imported,
    }


@mcp.tool()
def extract_text_from_pdf(
    project_id: str,
    pdf_path: str,
    max_chars: int = 50000,
) -> Dict[str, Any]:
    """
    Extract text from a PDF using pdfplumber and save it under the project.
    """
    pdf_file = Path(pdf_path).absolute()
    if not pdf_file.exists():
        raise ValueError("PDF not found")

    text = ""
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            text += page_text + "\n"

    extracted = text[:max_chars]
    out_path = project_path(project_id) / "parsed" / "tender_text.txt"
    safe_write(out_path, extracted)

    return {
        "status": "ok",
        "output": str(out_path),
        "preview": extracted[:1000],
    }


@mcp.tool()
def split_tender_by_section(project_id: str, tender_text: str) -> Dict[str, Any]:
    """
    Split tender text into key files: scope, evaluation, pricing, mandatory, terms.
    Heuristic only – the model should refine / check results.
    """
    root = project_path(project_id)
    section_dir = root / "sections_raw"
    ensure_dir(section_dir)

    lower = tender_text.lower()

    def extract(keyword: str, next_keywords: List[str]) -> str:
        idx = lower.find(keyword.lower())
        if idx == -1:
            return ""
        end = len(lower)
        for nk in next_keywords:
            j = lower.find(nk.lower(), idx + len(keyword))
            if j != -1 and j < end:
                end = j
        return tender_text[idx:end].strip()

    sections: Dict[str, str] = {
        "scope_of_work.txt": extract(
            "scope of work",
            ["evaluation", "pricing", "mandatory", "terms and conditions"],
        ),
        "evaluation_criteria.txt": extract(
            "evaluation",
            ["pricing", "mandatory", "terms and conditions"],
        ),
        "pricing_instructions.txt": extract(
            "pricing",
            ["mandatory", "terms and conditions"],
        ),
        "mandatory_requirements.txt": extract(
            "mandatory",
            ["terms and conditions"],
        ),
        "terms_and_conditions.txt": extract("terms and conditions", []),
    }

    out_files: Dict[str, str] = {}
    for name, content in sections.items():
        path = section_dir / name
        safe_write(path, content)
        out_files[name] = str(path)

    return {
        "status": "ok",
        "sections": out_files,
    }


# ============================================================
# SECTION B — COMPLIANCE TOOLS
# ============================================================

@mcp.tool()
def generate_compliance_checklist(project_id: str, tender_text: str) -> Dict[str, Any]:
    """
    Identify mandatory South African compliance items (CSD, SARS, PSIRA, B-BBEE, etc.).
    Very heuristic – the model must still verify against the full tender.
    """
    items = [
        ("CSD", "Central Supplier Database report"),
        ("TAX_CLEAR", "Tax Clearance / PIN"),
        ("BBBEE", "B-BBEE Certificate or Affidavit"),
        ("COIDA", "Letter of Good Standing"),
        ("PSIRA", "PSIRA Registration (security)"),
        ("CIDB", "CIDB Grading (construction)"),
        ("UIF", "UIF Registration"),
        ("SBD1", "SBD1 Form"),
        ("SBD4", "SBD4 Form"),
        ("SBD6.1", "SBD6.1 Form"),
        ("SBD8", "SBD8 Form"),
        ("SBD9", "SBD9 Form"),
    ]

    lower = tender_text.lower()
    detected: List[Dict[str, Any]] = []

    for code, desc in items:
        present = code.lower() in lower or desc.lower() in lower
        detected.append(
            {
                "code": code,
                "description": desc,
                "detected": present,
            }
        )

    out = project_path(project_id) / "compliance" / "checklist.json"
    safe_write(out, json.dumps(detected, indent=2))

    return {
        "status": "ok",
        "output": str(out),
        "items": detected,
    }


@mcp.tool()
def check_statutory_compliance(
    project_id: str,
    checklist_path: str,
    company_docs_folder: str,
) -> Dict[str, Any]:
    """
    Compare checklist vs company documents present in a folder (by filename heuristic).

    checklist_path: path to checklist.json from generate_compliance_checklist.
    company_docs_folder: folder containing company compliance PDFs (CSD, Tax, B-BBEE, PSIRA, etc.).
    """
    checklist_file = Path(checklist_path).absolute()
    if not checklist_file.exists():
        raise ValueError("Checklist file not found")

    folder = Path(company_docs_folder).absolute()
    if not folder.exists() or not folder.is_dir():
        raise ValueError("company_docs_folder must be an existing directory")

    checklist = json.loads(checklist_file.read_text(encoding="utf-8"))
    docs = [p.name.lower() for p in folder.iterdir() if p.is_file()]

    results: List[Dict[str, Any]] = []
    for item in checklist:
        code = item["code"].lower()
        match = any(code in d for d in docs)
        results.append(
            {
                "code": item["code"],
                "description": item["description"],
                "present": match,
            }
        )

    out = project_path(project_id) / "compliance" / "check_results.json"
    safe_write(out, json.dumps(results, indent=2))

    return {
        "status": "ok",
        "output": str(out),
        "results": results,
    }


@mcp.tool()
def validate_expiry_dates(
    project_id: str,
    doc_paths: List[str],
    max_chars_per_doc: int = 5000,
) -> Dict[str, Any]:
    """
    Scan PDFs for expiry-like dates and report them.
    Useful for B-BBEE, Tax, PSIRA certificates.
    """
    date_regex = re.compile(
        r"(20\d{2}[-/\.](0[1-9]|1[0-2])[-/\.](0[1-9]|[12]\d|3[01]))|"
        r"((0[1-9]|[12]\d|3[01])[-/\.](0[1-9]|1[0-2])[-/\.](20\d{2}))"
    )
    results: List[Dict[str, Any]] = []

    for p in doc_paths:
        fp = Path(p).absolute()
        if not fp.exists():
            results.append({"file": str(fp), "error": "not found"})
            continue
        if fp.suffix.lower() != ".pdf":
            results.append({"file": str(fp), "error": "only PDF supported"})
            continue

        text = ""
        with pdfplumber.open(fp) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                text += page_text + "\n"
                if len(text) >= max_chars_per_doc:
                    break
        text = text[:max_chars_per_doc]
        matches = date_regex.findall(text)
        flat_matches = {"".join(m) for m in matches}
        candidates = list(flat_matches)

        results.append(
            {
                "file": str(fp),
                "expiry_candidates": candidates[-5:],  # last few occurrences
            }
        )

    out = project_path(project_id) / "compliance" / "expiry_scan.json"
    safe_write(out, json.dumps(results, indent=2))

    return {
        "status": "ok",
        "output": str(out),
        "results": results,
    }


# ============================================================
# SECTION C — PRICING TOOLS
# ============================================================

@mcp.tool()
def parse_pricing_schedule(project_id: str, excel_path: str) -> Dict[str, Any]:
    """
    Parse an XLS/XLSX pricing schedule → JSON line items.

    Assumes the first sheet has columns:
    [Description, Unit, Qty, Rate, Total] (case-insensitive, best-effort).
    """
    excel_file = Path(excel_path).absolute()
    if not excel_file.exists():
        raise ValueError("Excel file not found")

    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb.active

    headers = [cell.value or "" for cell in sheet[1]]
    header_map = {str(h).strip().lower(): idx for idx, h in enumerate(headers)}

    def get_col(variants: List[str]) -> Optional[int]:
        for v in variants:
            idx = header_map.get(v.lower())
            if idx is not None:
                return idx
        return None

    desc_col = get_col(["description", "item", "activity"])
    unit_col = get_col(["unit", "uom"])
    qty_col = get_col(["qty", "quantity"])
    rate_col = get_col(["rate", "unit price", "unit_rate"])
    total_col = get_col(["total", "amount"])

    items: List[Dict[str, Any]] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue

        def val(col: Optional[int]) -> Any:
            if col is None or col >= len(row):
                return None
            return row[col]

        desc = val(desc_col) or ""
        unit = val(unit_col) or ""
        qty = val(qty_col) or 0
        rate = val(rate_col) or 0
        total = val(total_col)

        try:
            qty = float(qty or 0)
        except Exception:
            qty = 0.0
        try:
            rate = float(rate or 0)
        except Exception:
            rate = 0.0

        if total is None or total == "":
            total = qty * rate
        else:
            try:
                total = float(total)
            except Exception:
                total = qty * rate

        items.append(
            {
                "description": str(desc),
                "unit": str(unit),
                "qty": qty,
                "rate": rate,
                "total": total,
            }
        )

    out = project_path(project_id) / "pricing" / "items.json"
    safe_write(out, json.dumps(items, indent=2))

    return {
        "status": "ok",
        "items": items,
        "output": str(out),
    }


@mcp.tool()
def apply_markup_strategy(
    project_id: str,
    items: List[Dict[str, Any]],
    markup: float = 0.25,
    vat: float = 0.15,
) -> Dict[str, Any]:
    """
    Apply markup and VAT to line items.

    markup: 0.25 means 25% markup on base total.
    vat: 0.15 means 15% VAT on total_excl.
    """
    updated: List[Dict[str, Any]] = []
    for i in items:
        qty = float(i.get("qty", 0) or 0)
        rate = float(i.get("rate", 0) or 0)
        base = qty * rate
        excl = base * (1 + markup)
        vat_amt = excl * vat
        incl = excl + vat_amt
        updated.append(
            {
                **i,
                "base_total": round(base, 2),
                "total_excl": round(excl, 2),
                "vat": round(vat_amt, 2),
                "total_incl": round(incl, 2),
            }
        )

    out = project_path(project_id) / "pricing" / "priced.json"
    safe_write(out, json.dumps(updated, indent=2))

    summary = {
        "base_total": round(sum(i["base_total"] for i in updated), 2),
        "total_excl": round(sum(i["total_excl"] for i in updated), 2),
        "vat": round(sum(i["vat"] for i in updated), 2),
        "total_incl": round(sum(i["total_incl"] for i in updated), 2),
    }

    return {
        "status": "ok",
        "output": str(out),
        "items": updated,
        "summary": summary,
    }


@mcp.tool()
def build_pricing_excel(
    project_id: str,
    items: List[Dict[str, Any]],
    output_filename: str = "priced.xlsx",
) -> Dict[str, Any]:
    """
    Write priced items to an Excel file.

    Columns: Description, Unit, Qty, Rate, Base Total, Total Excl, VAT, Total Incl
    """
    root = project_path(project_id)
    out = root / "pricing" / output_filename
    ensure_dir(out.parent)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pricing"

    ws.append(
        [
            "Description",
            "Unit",
            "Qty",
            "Rate",
            "Base Total",
            "Total Excl",
            "VAT",
            "Total Incl",
        ]
    )

    for i in items:
        ws.append(
            [
                i.get("description", ""),
                i.get("unit", ""),
                i.get("qty", 0),
                i.get("rate", 0),
                i.get("base_total", 0),
                i.get("total_excl", 0),
                i.get("vat", 0),
                i.get("total_incl", 0),
            ]
        )

    wb.save(out)
    return {
        "status": "ok",
        "excel": str(out),
    }


@mcp.tool()
def sanity_check_pricing(
    project_id: str,
    items: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """
    Check arithmetic and flag suspiciously low/high values (rough heuristic).
    """
    warnings: List[str] = []
    checked: List[Dict[str, Any]] = []

    for idx, i in enumerate(items, start=1):
        qty = float(i.get("qty", 0) or 0)
        rate = float(i.get("rate", 0) or 0)
        expected_total = qty * rate
        given_total = float(i.get("total", expected_total) or expected_total)
        diff = abs(given_total - expected_total)
        if diff > 1:
            warnings.append(
                f"Row {idx}: total mismatch (expected {expected_total:.2f}, got {given_total:.2f})"
            )
        if rate < 10:
            warnings.append(f"Row {idx}: rate {rate} seems very low.")
        if rate > 1_000_000:
            warnings.append(f"Row {idx}: rate {rate} seems extremely high.")
        checked.append(
            {**i, "expected_total": expected_total, "total_diff": diff}
        )

    out = project_path(project_id) / "pricing" / "sanity_check.json"
    safe_write(
        out, json.dumps({"items": checked, "warnings": warnings}, indent=2)
    )

    return {
        "status": "ok",
        "output": str(out),
        "warnings": warnings,
    }


# ============================================================
# SECTION D — EVALUATION & RISK
# ============================================================

@mcp.tool()
def build_evaluation_matrix(project_id: str, criteria: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Build an evaluation matrix JSON from a list of criteria.

    criteria: list of {name, max_score, weight?, evidence_hint?}
    """
    root = project_path(project_id)
    out = root / "evaluation" / "matrix.json"
    safe_write(out, json.dumps(criteria, indent=2))
    return {"status": "ok", "output": str(out), "criteria": criteria}


@mcp.tool()
def self_score_response(
    project_id: str, scored_criteria: List[Dict[str, Any]]
) -> Dict[str, Any]:
    """
    Store a self-scored evaluation report.

    scored_criteria: list of {name, max_score, awarded_score, comments?}
    """
    total_max = sum(c.get("max_score", 0) for c in scored_criteria)
    total_awarded = sum(c.get("awarded_score", 0) for c in scored_criteria)
    percentage = (total_awarded / total_max * 100) if total_max else 0.0

    payload = {
        "scored_criteria": scored_criteria,
        "totals": {
            "total_max": total_max,
            "total_awarded": total_awarded,
            "percentage": percentage,
        },
    }

    out = project_path(project_id) / "evaluation" / "self_score.json"
    safe_write(out, json.dumps(payload, indent=2))

    return {
        "status": "ok",
        "output": str(out),
        "totals": payload["totals"],
    }


@mcp.tool()
def risk_register_builder(
    project_id: str,
    risks: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """
    Generate a risk register (JSON only).

    risks: list of {description, impact, likelihood, mitigation, owner?}
    """
    out = project_path(project_id) / "risk" / "risk_register.json"
    safe_write(out, json.dumps(risks, indent=2))
    return {
        "status": "ok",
        "output": str(out),
        "risks": risks,
    }


# ============================================================
# SECTION E — COMPANY PROFILE & EVIDENCE
# ============================================================

@mcp.tool()
def company_profile_from_json(project_id: str, profile: Dict[str, Any]) -> Dict[str, Any]:
    """
    Save a normalized company profile JSON.
    """
    out = project_path(project_id) / "company" / "profile.json"
    safe_write(out, json.dumps(profile, indent=2))
    return {
        "status": "ok",
        "output": str(out),
        "profile": profile,
    }


@mcp.tool()
def project_reference_pack_builder(
    project_id: str, references: List[Dict[str, Any]]
) -> Dict[str, Any]:
    """
    Save a list of past project references.

    references: list of {name, client, value?, duration?, description?}
    """
    out = project_path(project_id) / "company" / "references.json"
    safe_write(out, json.dumps(references, indent=2))
    return {"status": "ok", "output": str(out), "references": references}


@mcp.tool()
def cv_normalizer(
    project_id: str, staff: List[Dict[str, Any]]
) -> Dict[str, Any]:
    """
    Normalize staff CVs to a consistent structure.

    staff: list of {name, role, qualifications?, experienceSummary?, psiraGrade?}
    """
    out = project_path(project_id) / "staff" / "staff.json"
    safe_write(out, json.dumps(staff, indent=2))
    return {"status": "ok", "output": str(out), "staff": staff}


@mcp.tool()
def media_gallery_builder(
    project_id: str,
    image_paths: List[str],
    title: str = "Project Gallery",
) -> Dict[str, Any]:
    """
    Build a basic HTML gallery from image paths (used in proposals).
    """
    root = project_path(project_id)
    out = root / "sections" / "gallery.html"
    ensure_dir(out.parent)

    items_html = "\n".join(
        f'<div class="thumb"><img src="{p}" alt="image" /></div>'
        for p in image_paths
    )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<title>{title}</title>
<style>
body {{ font-family: system-ui, sans-serif; padding: 32px; }}
.grid {{ display: grid; grid-template-columns: repeat(auto-fill,minmax(180px,1fr)); gap: 16px; }}
.thumb img {{ width: 100%; height: auto; border-radius: 8px; object-fit: cover; }}
</style>
</head>
<body>
<h1>{title}</h1>
<div class="grid">
{items_html}
</div>
</body>
</html>
"""
    safe_write(out, html)
    return {"status": "ok", "output": str(out), "images": image_paths}


# ============================================================
# SECTION F — PROPOSAL ASSEMBLY & EXPORT
# ============================================================

@mcp.tool()
def build_styled_section(
    project_id: str,
    section_name: str,
    heading: str,
    body_html: str,
) -> Dict[str, Any]:
    """
    Wrap body HTML in a styled section wrapper. Creates sections/{section_name}.html
    """
    root = project_path(project_id)
    out = root / "sections" / f"{section_name}.html"
    ensure_dir(out.parent)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<title>{heading}</title>
<style>
body {{ font-family: system-ui, sans-serif; padding: 32px; line-height: 1.5; }}
h1 {{ margin-bottom: 12px; }}
</style>
</head>
<body>
<h1>{heading}</h1>
{body_html}
</body>
</html>
"""
    safe_write(out, html)
    return {"status": "ok", "output": str(out)}


@mcp.tool()
def generate_cover_page(
    project_id: str,
    tender_number: str,
    tender_title: str,
    client_name: str,
    company_name: str,
) -> Dict[str, Any]:
    """
    Generate a simple cover page HTML.
    """
    root = project_path(project_id)
    out = root / "sections" / "cover_page.html"
    ensure_dir(out.parent)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<title>Cover Page</title>
<style>
body {{
  font-family: system-ui, sans-serif;
  padding: 64px;
  text-align: center;
}}
h1 {{ font-size: 32px; margin-bottom: 8px; }}
h2 {{ font-size: 20px; margin-bottom: 24px; }}
p  {{ font-size: 14px; margin-bottom: 6px; }}
</style>
</head>
<body>
<h1>{company_name}</h1>
<h2>Tender Response</h2>
<p><strong>Tender No:</strong> {tender_number}</p>
<p><strong>Tender Title:</strong> {tender_title}</p>
<p><strong>Client:</strong> {client_name}</p>
</body>
</html>
"""
    safe_write(out, html)
    return {"status": "ok", "output": str(out)}


@mcp.tool()
def generate_table_of_contents(
    project_id: str,
    entries: List[str],
) -> Dict[str, Any]:
    """
    Generate a basic HTML table of contents.
    """
    root = project_path(project_id)
    out = root / "sections" / "table_of_contents.html"
    ensure_dir(out.parent)

    list_items = "\n".join(f"<li>{e}</li>" for e in entries)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<title>Table of Contents</title>
<style>
body {{ font-family: system-ui, sans-serif; padding: 32px; }}
ol {{ margin-left: 16px; }}
</style>
</head>
<body>
<h1>Table of Contents</h1>
<ol>
{list_items}
</ol>
</body>
</html>
"""
    safe_write(out, html)
    return {"status": "ok", "output": str(out)}


@mcp.tool()
def assemble_document(
    project_id: str,
    section_paths: List[str],
    output_filename: str = "proposal.html",
) -> Dict[str, Any]:
    """
    Merge multiple HTML fragments into a single HTML document in order.
    """
    root = project_path(project_id)
    out = root / output_filename

    combined_body = ""
    for p in section_paths:
        fp = Path(p).absolute()
        if fp.exists():
            combined_body += read_text_file(fp) + "\n\n"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<title>Tri-Tender Proposal</title>
</head>
<body>
{combined_body}
</body>
</html>
"""
    safe_write(out, html)
    return {"status": "ok", "output": str(out)}


@mcp.tool()
def html_to_docx(
    project_id: str,
    html_path: str,
    output_filename: str = "proposal.docx",
) -> Dict[str, Any]:
    """
    VERY BASIC: Convert HTML text to a DOCX where each line becomes a paragraph.

    This is not a full HTML → DOCX conversion, but enough for editable drafts.
    Requires python-docx.
    """
    if docx is None:
        raise RuntimeError("python-docx is not installed")

    root = project_path(project_id)
    out = root / output_filename
    ensure_dir(out.parent)

    html_text = read_text_file(Path(html_path).absolute())
    # Strip simple tags and split on newline
    text = re.sub(r"<[^>]+>", "", html_text)
    lines = [l.strip() for l in text.splitlines() if l.strip()]

    document = docx.Document()
    for line in lines:
        document.add_paragraph(line)
    document.save(out)

    return {"status": "ok", "output": str(out)}


@mcp.tool()
def html_to_pdf(
    project_id: str,
    html_path: str,
    output_filename: str = "proposal.pdf",
) -> Dict[str, Any]:
    """
    Render HTML to a real PDF using WeasyPrint (if available).

    On Windows, WeasyPrint also needs system libraries (Cairo, Pango, etc.).
    If those are missing, this tool will raise a clear error, but the MCP
    server will still start and all other tools will work.
    """
    try:
        from weasyprint import HTML as WeasyHTML_local
    except Exception as e:
        raise RuntimeError(
            "WeasyPrint is not available or its system libraries are missing. "
            "Install it following the official docs before using html_to_pdf."
        ) from e

    root = project_path(project_id)
    out = root / output_filename
    ensure_dir(out.parent)

    html_file = Path(html_path).absolute()
    if not html_file.exists():
        raise ValueError(f"HTML file not found: {html_file}")

    WeasyHTML_local(filename=str(html_file)).write_pdf(str(out))

    return {
        "status": "ok",
        "output": str(out),
        "note": "Rendered with WeasyPrint.",
    }


# ============================================================
# SECTION G — WORKFLOW, DEADLINES & PACKAGING
# ============================================================

@mcp.tool()
def create_workflow_log_entry(
    project_id: str,
    agent_name: str,
    action: str,
    notes: str = "",
    timestamp: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Append a workflow log entry (JSONL).
    """
    root = project_path(project_id)
    log_path = root / "workflow_log.jsonl"
    entry = {
        "timestamp": timestamp or datetime.utcnow().isoformat() + "Z",
        "agent_name": agent_name,
        "action": action,
        "notes": notes,
    }
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry) + "\n")

    return {"status": "ok", "log_path": str(log_path), "entry": entry}


@mcp.tool()
def generate_project_checklist(
    project_id: str,
    items: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """
    Store a project checklist JSON.

    items: list of {id, label, category?}
    """
    root = project_path(project_id)
    out = root / "checklist" / "checklist.json"
    ensure_dir(out.parent)
    safe_write(out, json.dumps(items, indent=2))
    return {"status": "ok", "output": str(out), "items": items}


@mcp.tool()
def deadline_tracker(
    project_id: str,
    closing_date_time: str,
    milestones: Optional[List[Dict[str, str]]] = None,
    write_ics: bool = True,
) -> Dict[str, Any]:
    """
    Store tender closing date and milestones; optionally write an .ics file.
    """
    if milestones is None:
        milestones = []

    root = project_path(project_id)
    deadlines_json = root / "deadlines.json"
    payload = {"closing_date_time": closing_date_time, "milestones": milestones}
    safe_write(deadlines_json, json.dumps(payload, indent=2))

    ics_path = None
    if write_ics:
        def iso_to_ics(iso: str) -> str:
            # "2025-11-30T11:00:00" → "20251130T110000Z"
            clean = iso.replace("-", "").replace(":", "").replace(" ", "")
            if "." in clean:
                clean = clean.split(".")[0]
            clean = clean.replace("T", "")
            if not clean.endswith("Z"):
                clean += "Z"
            return clean

        dtstamp = iso_to_ics(datetime.utcnow().isoformat())

        lines = [
            "BEGIN:VCALENDAR",
            "VERSION:2.0",
            "PRODID:-//TriTender//DeadlineTracker//EN",
        ]
        # Closing event
        lines.extend(
            [
                "BEGIN:VEVENT",
                f"UID:{project_id}-closing@tri-tender",
                f"DTSTAMP:{dtstamp}",
                f"DTSTART:{iso_to_ics(closing_date_time)}",
                f"SUMMARY:Tender closing: {project_id}",
                "END:VEVENT",
            ]
        )
        # Milestone events
        for idx, m in enumerate(milestones, start=1):
            lines.extend(
                [
                    "BEGIN:VEVENT",
                    f"UID:{project_id}-m{idx}@tri-tender",
                    f"DTSTAMP:{dtstamp}",
                    f"DTSTART:{iso_to_ics(m['dueDateTime'])}",
                    f"SUMMARY:{project_id}: {m['name']}",
                    "END:VEVENT",
                ]
            )
        lines.append("END:VCALENDAR")

        ics_path = root / "deadlines.ics"
        ics_path.write_text("\r\n".join(lines), encoding="utf-8")

    return {
        "status": "ok",
        "deadlines_json": str(deadlines_json),
        "ics_path": str(ics_path) if ics_path else None,
    }


@mcp.tool()
def generate_submission_cover_letter(
    project_id: str,
    tender_number: str,
    client_name: str,
    company_name: str,
    brief_summary: str,
) -> Dict[str, Any]:
    """
    Generate a simple submission cover letter HTML.
    """
    root = project_path(project_id)
    out = root / "sections" / "cover_letter.html"
    ensure_dir(out.parent)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<title>Submission Cover Letter</title>
<style>
body {{ font-family: system-ui, sans-serif; padding: 48px; line-height: 1.5; }}
p {{ margin-bottom: 12px; }}
</style>
</head>
<body>
<p>{client_name}</p>
<p>Attention: Tender Committee</p>
<p>Bid/Tender No: {tender_number}</p>
<br />
<p><strong>Re: Submission of Tender – {tender_number}</strong></p>
<p>We, <strong>{company_name}</strong>, hereby submit our tender response for the above-mentioned bid.</p>
<p>{brief_summary}</p>
<p>We confirm that the required forms, declarations and supporting documents are attached as per the tender requirements.</p>
<p>We trust that our proposal will receive your favourable consideration.</p>
<br />
<p>Yours faithfully,</p>
<p><strong>{company_name}</strong></p>
</body>
</html>
"""
    safe_write(out, html)
    return {"status": "ok", "output": str(out)}


@mcp.tool()
def submission_package_builder(
    project_id: str,
    source_folders: List[str],
    output_folder: str = "submission",
    zip_name: str = "submission.zip",
) -> Dict[str, Any]:
    """
    Create a submission folder under the project, copy selected folders into it, and zip it.
    """
    root = project_path(project_id)
    submission_root = root / output_folder
    ensure_dir(submission_root)

    def copy_tree(src: Path, dst: Path):
        if not src.exists():
            return
        for item in src.rglob("*"):
            rel = item.relative_to(src)
            target = dst / rel
            if item.is_dir():
                ensure_dir(target)
            else:
                ensure_dir(target.parent)
                shutil.copy(item, target)

    for folder in source_folders:
        src = root / folder
        copy_tree(src, submission_root / folder)

    zip_path = root / zip_name
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for file in submission_root.rglob("*"):
            zf.write(file, file.relative_to(root))

    return {
        "status": "ok",
        "submission_folder": str(submission_root),
        "zip_path": str(zip_path),
    }


# ============================================================
# RUNNER
# ============================================================

if __name__ == "__main__":
    # For local dev with Claude Desktop & MCP Inspector:
    mcp.run()
